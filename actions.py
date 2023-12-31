from flask import Blueprint, request, render_template, redirect, url_for, flash, Response, jsonify
import re
import os
import xml.dom.minidom
import logging
import openpyxl
from datetime import datetime, timedelta
import warnings
import pandas as pd
import configparser
import subprocess

from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter

# Cấu hình logging
logging.basicConfig(filename='logfile.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

logging.info('#################################################')
logging.info('#                 Start APP                     #')
logging.info('#################################################')

actions = Blueprint('actions', __name__, template_folder='templates')

parser = configparser.ConfigParser()


def count_leading_spaces(line):
    """Đếm số khoảng trắng ở đầu mỗi dòng."""
    count = 0
    for char in line:
        if char == ' ':
            count += 1
        elif char == '\t':
            count += 4
        else:
            break  # Ngừng khi gặp ký tự không phải khoảng trắng
    return count


@actions.route('/convert-project', methods=['POST', 'GET'])
def convert_project():
    parser.read("./config/config.txt", encoding="utf-8")
    project_path = parser.get("config", "project_path")
    pattern_path = parser.get("config", "pattern_path")
    return render_template('convert-project.html', project_path=project_path, pattern_path=pattern_path)


@actions.route('/convert-file', methods=['POST', 'GET'])
def convert_file():
    return render_template('convert-file.html')


@actions.route('/settings', methods=['POST', 'GET'])
def setting_file():
    config_path = os.path.abspath(os.path.dirname(__file__)) + "\config\config.txt"
    return render_template('settings.html', config_path=config_path)


@actions.route('/open-config', methods=['POST'])
def open_setting_file():
    try:
        file_path = os.path.abspath(os.path.dirname(__file__)) + "\config\config.txt"
        subprocess.Popen(['notepad.exe', file_path])
        return render_template('settings.html', config_path=file_path)
    except Exception as e:
        # Xảy ra lỗi khi kết nối
        print('Error: ' + str(e))
        logging.info("ERROR: " + str(e))
        return jsonify({'status': 'error', 'message': str(e)})


@actions.route('/', methods=['POST', 'GET'])
def convert_code():
    return render_template('convert-code.html')


@actions.route('/start-convert-files', methods=['POST', 'GET'])
def start_convert_files():
    try:
        logging.info("#### [start-convert-files] ####")
        parser.read("./config/config.txt", encoding="utf-8")
        pattern_path = parser.get("config", "pattern_path")
        conversion_rules_file_path = pattern_path
        conversion_rules = read_conversion_rules_from_file_excel(conversion_rules_file_path)

        isExist = os.path.exists("Output_file")
        if not isExist:
            os.makedirs("Output_file")
        if request.method == 'POST':
            files = request.files.getlist('filepond')
            workbook = openpyxl.Workbook()
            for file in files:
                item_sheet = workbook.create_sheet(title=file.filename)

                # Tạo lại cấu trúc thư mục gốc
                file_type = file.filename.split('.')[-1]
                rule_for_file_type = []
                for category, pattern, regex, replace, pic, filetype, flagList in conversion_rules:
                    if filetype != None:
                        if file_type.strip() == filetype.strip():
                            data = [category, pattern, regex, replace, pic, filetype, flagList]
                            rule_for_file_type.append(data)

                data_result = process_code_file(file, rule_for_file_type)
                file_convert = "Output_file/" + file.filename
                with open(file_convert, "w", encoding="utf-8", newline='\n') as sql_file:
                    sql_file.write(data_result)

        return jsonify({'status': 'success', 'message': 'Convert successful!'})
    except Exception as e:
        # Xảy ra lỗi khi kết nối
        print('Error: ' + str(e))
        logging.info("ERROR: " + str(e))
        return jsonify({'status': 'error', 'message': str(e)})


@actions.route('/convert-editor', methods=['POST'])
def convert_editor():
    try:
        print("### [convert_editor] ###")
        today = datetime.now().strftime("%Y/%m/%d")
        from_code = request.form['code-editor-1']
        parser.read("./config/config.txt", encoding="utf-8")
        pattern_path = parser.get("config", "pattern_path")
        conversion_rules_file_path = pattern_path
        conversion_rules = read_conversion_rules_from_file_excel(conversion_rules_file_path)
        input_code = from_code
        output_code = input_code
        content_comment = parser.get("config", "content_comment")
        content_comment_del = parser.get("config", "content_comment_del")
        source_code = from_code
        for category, pattern, regex, replace, pic, filetype, flagList in conversion_rules:
            if (pattern != None) and (regex != None) and (regex.strip() != "TBD"):
                match_pattern = re.findall(regex, source_code, flags=re.IGNORECASE)
                if match_pattern:
                    match_pattern = list(set(match_pattern))
                    if (replace == None):
                        flag_check = 1
                        # DELETE CODE COMMENT
                    else:
                        flag_check = 1

                        for index, match_data in enumerate(match_pattern):
                            count_space = count_leading_spaces(match_data[0])
                            old_code = ''.join(map(str, match_data))
                            new_code = re.sub(regex.strip(), replace, old_code, flags=re.IGNORECASE)
                            if flagList != True:
                                if (filetype.lower() == "jsp"):
                                    source_code = source_code.replace(old_code,
                                                                      count_space * ' ' + '<%-- (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + ' --%>' + '\n' +
                                                                      count_space * ' ' + '<%-- ' + old_code.strip() + ' --%>' + '\n' +
                                                                      count_space * ' ' + new_code.strip() + '\n' +
                                                                      count_space * ' ' + '<%-- (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '--%>'
                                                                      )
                                elif (filetype.lower() == "html") or (filetype.lower() == "xml"):
                                    source_code = source_code.replace(old_code,
                                                                      count_space * ' ' + '<!-- (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + ' -->' + '\n' +
                                                                      count_space * ' ' + '<!-- ' + old_code.strip() + ' -->' + '\n' +
                                                                      count_space * ' ' + new_code.strip() + '\n' +
                                                                      count_space * ' ' + '<!-- (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '-->'
                                                                      )
                                else:
                                    source_code = source_code.replace(old_code,
                                                                      count_space * ' ' + '// (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '\n' +
                                                                      count_space * ' ' + '// ' + old_code.strip() + '\n' +
                                                                      count_space * ' ' + new_code.strip() + '\n' +
                                                                      count_space * ' ' + '// (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category
                                                                      )
                            else:
                                source_code = source_code.replace(old_code, new_code)

        to_code = source_code
        # for category, pattern, regex, replace, pic in conversion_rules:
        #     if (pattern != None) and (regex != None) and (regex.strip() != "TBD"):
        #         match_pattern = re.findall(regex, pattern, flags=re.IGNORECASE)
        #         if match_pattern:
        #             if (replace == None):
        #                 output_code = re.sub(regex,
        #                                      r'//(STR) ' + today + ' ' + content_comment_del + ' ' + pic + ' DEL ' + category + r'\n' + '//' + pattern + r'\n' + '//(END) ' + today + ' ' + content_comment_del + ' ' + pic + ' MOD ' + category,
        #                                      output_code, flags=re.IGNORECASE)
        #             else:
        #                 # MODIFY CODE COMMENT
        #                 output_code = re.sub(regex,
        #                                      r'//(STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + r'\n' + '//' + pattern + r'\n' + replace + r'\n' + '//(END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category,
        #                                      output_code, flags=re.IGNORECASE)
        #to_code = output_code

        return jsonify({
            "from_code": from_code,
            "to_code": to_code
        })

    except Exception as e:
        print('Error: ' + str(e))
        logging.info("[convert-editor] ERROR: " + str(e))
        return jsonify({'status': 'error', 'message': str(e)})


def cell_in_merged_range(cell, merged_ranges):
    for range in merged_ranges:
        if range.contains(cell.row, cell.column):
            return True
    return False


def read_conversion_rules_from_file_excel(file_path):
    try:
        # Mở workbook
        workbook = openpyxl.load_workbook(file_path)
        # Chọn sheet cần đọc
        sheet = workbook['Pattern Summary']

        # Lấy tên cột từ hàng thứ 4 (index 4)
        column_names = [cell.value for cell in sheet[4]]

        # Đọc dữ liệu từ hàng thứ 5 trở đi
        data = sheet.iter_rows(min_row=5, values_only=True)
        df = pd.DataFrame(data, columns=column_names)

        merged_cells = sheet.merged_cells.ranges
        # Chuyển tập hợp thành danh sách các tuple

        merged_cells_list = [str(cell) for cell in merged_cells]
        column_letters = [get_column_letter(col + 1) for col in range(len(column_names))]
        # Lặp qua từng ô merge
        i = 0
        for column in column_names:
            letter = column_letters[i]
            for merged_cell in merged_cells_list:
                start, end = merged_cell.split(":")
                start_row = int(start[1:]) - 5
                end_row = int(end[1:]) - 5
                start_col = get_column_letter(sheet[start].column)
                if (start_col == letter):
                    value_merge = df.at[start_row, column]
                    for row in range(start_row, end_row + 1):
                        df.at[row, column] = value_merge
            i = i + 1

        # Lấy ra các giá trị từ cột "Category"
        category_list = df['Category'].tolist()
        # Lấy ra các giá trị từ cột "Pattern"
        pattern_list = df['Pattern'].tolist()
        # Lấy ra các giá trị từ cột "Regex"
        regex_list = df['Regex'].tolist()
        # Lấy ra các giá trị từ cột "Replace"
        replace_list_1 = df['Replace '].tolist()
        replace_list = []
        # Lấy ra các giá trị từ cột "PIC"
        pic_list = df['PIC'].tolist()
        # Lấy ra các giá trị từ cột "FileType"
        file_list = df['FileType'].tolist()
        # Lấy ra các giá trị từ cột "NoCommentFlag"
        no_comment_flag_lst = df['NoCommentFlag'].tolist()

        for item in replace_list_1:
            if item != None and item.strip() != "TBD":
                item = re.sub(r'\$', '\\\\', item, flags=re.IGNORECASE)
            replace_list.append(item)
        return list(
            zip(category_list, pattern_list, regex_list, replace_list, pic_list, file_list, no_comment_flag_lst))
    except Exception as e:
        # Xảy ra lỗi khi kết nối
        print('Error: ' + str(e))
        logging.info("[read_conversion_rules_from_file_excel] ERROR: " + str(e))
        return jsonify({'status': 'error', 'message': str(e)})


@actions.route('/start-convert-folder', methods=['POST'])
def convert_folder():
    try:
        files = []
        files_path = []
        today = datetime.now().strftime("%Y%m%d")
        # Lấy giờ, phút, giây hiện tại
        current_time = datetime.now().strftime("%H%M%S")
        # Tạo tên thư mục với ngày và giờ phút giây
        folder_name = f"{today}_{current_time}"

        save_dir = os.path.join('output_source_folder', folder_name)

        parser.read("./config/config.txt", encoding="utf-8")
        # pattern_path = parser.get("config", "pattern_path")
        pattern_path = request.form['patternPath']

        # project_path = parser.get("config", "project_path")
        project_path = request.form['folderPath']

        conversion_rules_file_path = pattern_path
        conversion_rules = read_conversion_rules_from_file_excel(conversion_rules_file_path)

        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        for root, dirnames, filenames in os.walk(project_path):
            if root == project_path:
                save_root = save_dir
            else:
                rel_dir = os.path.relpath(root, project_path)
                save_root = os.path.join(save_dir, rel_dir)
            if not os.path.exists(save_root):
                os.makedirs(save_root)

            for filename in filenames:
                if filename.endswith('.java') or filename.endswith('.jsp') or filename.endswith(
                        '.html') or filename.endswith('.js') or filename.endswith('.xml'):
                    filepath = os.path.join(root, filename)
                    # Tạo lại cấu trúc thư mục gốc
                    file_type = filename.rsplit('.', 1)[-1]
                    save_path = os.path.join(save_root, filename)
                    rule_for_file_type = []
                    for category, pattern, regex, replace, pic, filetype, flagList in conversion_rules:
                        if filetype != None:
                            if file_type.strip() == filetype.strip():
                                data = [category, pattern, regex, replace, pic, filetype, flagList]
                                rule_for_file_type.append(data)

                    data_result = process_code_folder(filepath, rule_for_file_type)
                    if data_result != "":
                        with open(save_path, 'w', encoding="utf-8") as f:
                            f.write(data_result)

                        files.append(filename)
                        files_path.append(filepath)
        return jsonify({'status': 'success', 'message': 'Convert successful!'})
    except Exception as e:
        print('Error: ' + str(e))
        logging.info("[start-convert-folder] ERROR: " + str(e))
        return jsonify({'status': 'error', 'message': str(e)})


def convert_with_pattern_file_excel_by_lines(source_code, conversion_rules):
    pattern = ''
    replace = ''
    regex = ''
    flag_check = 0  # khong sua gì file
    try:
        parser.read("./config/config.txt", encoding="utf-8")
        content_header = parser.get("config", "content_header")
        content_comment = parser.get("config", "content_comment")
        content_comment_del = parser.get("config", "content_comment_del")
        today = datetime.now().strftime("%Y/%m/%d")
        today_his = datetime.now().strftime("%Y.%m.%d")
        output_code = []
        pattern_history = r'修正履歴：XXXX.XX.XX XXX-XXX Name'
        pattern_history2 = r'修正履歴：20XX.XX.XX'
        pattern_history3 = r'修正履歴：200X.XX.XX XXX-XXX Name'
        pattern_history4 = r'修正履歴.*(\d{4})(-|\/|.)(\d{2})(-|\/|.)(\d{2})([^\n]*)'
        for line in source_code:
            count_space = count_leading_spaces(line)
            for category, pattern, regex, replace, pic, filetype, flagList in conversion_rules:
                if (pattern != None) and (regex != None) and (regex.strip() != "TBD"):
                    match_pattern = re.findall(regex, line, flags=re.IGNORECASE)
                    if match_pattern:
                        if (replace == None):
                            flag_check = 1
                            # DELETE CODE COMMENT
                            if (filetype.lower() == "html") or (filetype.lower() == "xml"):
                                line = re.sub(regex,
                                              r'<!-- (STR) ' + today + ' ' + content_comment_del + ' ' + pic + ' DEL ' + category + ' -->' + r'\n' + count_space * ' ' + '<!--' + line.strip() + '-->' + r'\n' + count_space * ' ' + '<!-- (END) ' + today + ' ' + content_comment_del + ' ' + pic + ' DEL ' + category + '-->',
                                              line, flags=re.IGNORECASE)
                            elif (filetype.lower() == "jsp"):
                                line = re.sub(regex,
                                              r'<%-- (STR) ' + today + ' ' + content_comment_del + ' ' + pic + ' DEL ' + category + ' --%>' + r'\n' + count_space * ' ' + '<%--' + line.strip() + '--%>' + r'\n' + count_space * ' ' + '<%-- (END) ' + today + ' ' + content_comment_del + ' ' + pic + ' DEL ' + category + '--%>',
                                              line, flags=re.IGNORECASE)
                            else:
                                line = re.sub(regex,
                                              r'//(STR) ' + today + ' ' + content_comment_del + ' ' + pic + ' DEL ' + category + r'\n' + count_space * ' ' + '//' + line.strip() + r'\n' + count_space * ' ' + '//(END) ' + today + ' ' + content_comment_del + ' ' + pic + ' DEL ' + category,
                                              line, flags=re.IGNORECASE)
                        else:
                            flag_check = 1
                            if flagList != True:
                                # MODIFY CODE COMMENT
                                if (filetype.lower() == "html") or (filetype.lower() == "xml"):
                                    line = line.strip().replace(r'\t', '')
                                    line = re.sub(regex,
                                                  count_space * ' ' + r'<!-- (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + ' -->' + r'\n' + '<!-- ' + line.strip() + ' -->' + r'\n' + count_space * ' ' + replace + r'\n' + count_space * ' ' + '<!-- (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '-->',
                                                  line, flags=re.IGNORECASE)
                                elif (filetype.lower() == "jsp"):
                                    line = line.strip().replace(r'\t', '')
                                    line = re.sub(regex,
                                                  count_space * ' ' + r'<%-- (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + ' --%>' + r'\n' + count_space * ' ' + '<%-- ' + line.strip() + ' --%>' + r'\n' + count_space * ' ' + replace + r'\n' + count_space * ' ' + '<%-- (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '--%>',
                                                  line, flags=re.IGNORECASE)
                                else:
                                    flag_check = 1
                                    line = line.strip().replace(r'\t', '')
                                    line = re.sub(regex,
                                                  count_space * ' ' + r'// (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + r'\n' + count_space * ' ' + '// ' + line.strip() + r'\n' + count_space * ' ' + replace + r'\n' + count_space * ' ' + '// (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category,
                                                  line, flags=re.IGNORECASE)
                            else:
                                line = line.strip().replace(r'\t', '')
                                line = re.sub(regex,
                                              count_space * ' ' + replace,
                                              line, flags=re.IGNORECASE)

            output_code.append(line)

        # add history:
        if flag_check == 1:
            output_code_new = []
            for line_2 in output_code:
                match_pattern_his = re.findall(pattern_history, line_2, flags=re.IGNORECASE)
                match_pattern_his2 = re.findall(pattern_history2, line_2, flags=re.IGNORECASE)
                match_pattern_his3 = re.findall(pattern_history3, line_2, flags=re.IGNORECASE)

                if (match_pattern_his2):
                    flag_check = 2
                    line_2 = re.sub(pattern_history2,
                                    r'修正履歴：' + today_his + ' ' + content_header + r'\n' + r' * 修正履歴：20XX.XX.XX',
                                    line_2, flags=re.IGNORECASE)
                elif (match_pattern_his):
                    flag_check = 2
                    line_2 = re.sub(pattern_history,
                                    r'修正履歴：' + today_his + ' ' + content_header + r'\n' + r' * 修正履歴：XXXX.XX.XX XXX-XXX Name',
                                    line_2, flags=re.IGNORECASE)
                elif (match_pattern_his3):
                    flag_check = 2
                    line_2 = re.sub(pattern_history3,
                                    r'修正履歴：' + today_his + ' ' + content_header + r'\n' + r' * 修正履歴：200X.XX.XX XXX-XXX Name',
                                    line_2, flags=re.IGNORECASE)
                output_code_new.append(line_2)

            output_code = output_code_new

            result = ""
            last_index = len(output_code) - 1
            for index, line in enumerate(output_code):
                if index != last_index:
                    result += line + "\n"
                else:
                    result += line
            if flag_check == 1:
                match_pattern_his4 = re.findall(pattern_history4, result, flags=re.IGNORECASE)
                last_index = len(match_pattern_his4) - 1
                if (match_pattern_his4):
                    for index1, line1 in enumerate(match_pattern_his4):
                        if index1 == last_index:
                            line_temp = line1[0] + line1[1] + line1[2] + line1[3] + line1[4] + line1[5]
                            result = result.replace(line_temp,
                                                    line_temp + '\n' + r' * 修正履歴：' + today_his + ' ' + content_header)
        else:
            result = ""

        return result
    except Exception as e:
        print('Error: ' + str(e))
        logging.info("[convert_with_pattern_file_excel_by_lines] ERROR: " + str(e))
        logging.info("## pattern ERROR:  - pattern = " + str(pattern))
        logging.info("##                 - regex   = " + str(regex))
        logging.info("##                 - replace = " + str(replace))
        return jsonify({'status': 'error', 'message': str(e)})


def convert_with_pattern_file_excel_all(source_code, conversion_rules):
    pattern = ''
    replace = ''
    regex = ''
    flag_check = 0  # khong sua gì file
    try:
        parser.read("./config/config.txt", encoding="utf-8")
        content_header = parser.get("config", "content_header")
        content_comment = parser.get("config", "content_comment")
        content_comment_del = parser.get("config", "content_comment_del")
        today = datetime.now().strftime("%Y/%m/%d")
        today_his = datetime.now().strftime("%Y.%m.%d")
        output_code = []
        pattern_history = r'修正履歴：XXXX.XX.XX XXX-XXX Name'
        pattern_history2 = r'修正履歴：20XX.XX.XX'
        pattern_history3 = r'修正履歴：200X.XX.XX XXX-XXX Name'
        pattern_history4 = r'修正履歴.*(\d{4})(-|\/|.)(\d{2})(-|\/|.)(\d{2})([^\n]*)'
        for category, pattern, regex, replace, pic, filetype, flagList in conversion_rules:
            if (pattern != None) and (regex != None) and (regex.strip() != "TBD"):
                match_pattern = re.findall(regex, source_code, flags=re.IGNORECASE)
                if match_pattern:
                    match_pattern = list(set(match_pattern))
                    if (replace == None):
                        flag_check = 1
                        # DELETE CODE COMMENT
                    else:
                        flag_check = 1
                        for index, match_data in enumerate(match_pattern):
                            count_space = count_leading_spaces(match_data[0])
                            old_code = ''.join(map(str, match_data))
                            new_code = re.sub(regex.strip(), replace, old_code, flags=re.IGNORECASE)
                            if flagList != True:
                                if (filetype.lower() == "jsp"):
                                    source_code = source_code.replace(old_code,
                                                                      count_space * ' ' + '<%-- (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + ' --%>' + '\n' +
                                                                      count_space * ' ' + '<%-- ' + old_code.strip() + ' --%>' + '\n' +
                                                                      count_space * ' ' + new_code.strip() + '\n' +
                                                                      count_space * ' ' + '<%-- (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '--%>'
                                                                      )
                                elif (filetype.lower() == "html") or (filetype.lower() == "xml"):
                                    source_code = source_code.replace(old_code,
                                                                      count_space * ' ' + '<!-- (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + ' -->' + '\n' +
                                                                      count_space * ' ' + '<!-- ' + old_code.strip() + ' -->' + '\n' +
                                                                      count_space * ' ' + new_code.strip() + '\n' +
                                                                      count_space * ' ' + '<!-- (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '-->'
                                                                      )
                                else:
                                    source_code = source_code.replace(old_code,
                                                                      count_space * ' ' + '// (STR) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category + '\n' +
                                                                      count_space * ' ' + '// ' + old_code.strip() + '\n' +
                                                                      count_space * ' ' + new_code.strip() + '\n' +
                                                                      count_space * ' ' + '// (END) ' + today + ' ' + content_comment + ' ' + pic + ' MOD ' + category
                                                                      )
                            else:
                                source_code = source_code.replace(old_code, new_code)
        # add history:
        if flag_check == 1:
            match_pattern_his = re.findall(pattern_history, source_code, flags=re.IGNORECASE)
            match_pattern_his2 = re.findall(pattern_history2, source_code, flags=re.IGNORECASE)
            match_pattern_his3 = re.findall(pattern_history3, source_code, flags=re.IGNORECASE)
            match_pattern_his4 = re.findall(pattern_history4, source_code, flags=re.IGNORECASE)
            if (match_pattern_his):
                last_index_his = len(match_pattern_his) - 1
                old_his = ''.join(map(str, match_pattern_his[last_index_his]))
                new_his = old_his.replace(old_his, '修正履歴：' + str(today_his) + ' ' + content_header + '\n' + ' * 修正履歴：XXXX.XX.XX XXX-XXX Name')
                source_code = source_code.replace(old_his, new_his)
            elif (match_pattern_his2):
                last_index_his2 = len(match_pattern_his2) - 1
                old_his = ''.join(map(str, match_pattern_his2[last_index_his2]))
                new_his = old_his.replace(old_his, '修正履歴：' + today_his + ' ' + content_header + '\n' + ' * 修正履歴：20XX.XX.XX')
                source_code = source_code.replace(old_his, new_his)
            elif (match_pattern_his3):
                last_index_his3 = len(match_pattern_his3) - 1
                old_his = ''.join(map(str, match_pattern_his3[last_index_his3]))
                new_his = old_his.replace(old_his, '修正履歴：' + today_his + ' ' + content_header + '\n' + ' * 修正履歴：200X.XX.XX XXX-XXX Name')
                source_code = source_code.replace(old_his, new_his)
            elif (match_pattern_his4):
                last_index_his4 = len(match_pattern_his4) - 1
                old_his = ''.join(map(str, match_pattern_his4[last_index_his4]))
                source_code = source_code.replace(old_his,
                                        old_his + '\n' + r' * 修正履歴：' + today_his + ' ' + content_header)
        else:
            source_code = ""

        return source_code
    except Exception as e:
        print('Error: ' + str(e))
        logging.info("[convert_with_pattern_file_excel_by_lines] ERROR: " + str(e))
        logging.info("## pattern ERROR:  - pattern = " + str(pattern))
        logging.info("##                 - regex   = " + str(regex))
        logging.info("##                 - replace = " + str(replace))
        return jsonify({'status': 'error', 'message': str(e)})


def process_code_folder(file, conversion_rules):
    try:
        with open(file, encoding="utf-8") as f:
            source_code = f.read()
        if isinstance(source_code, bytes):
            source_code = source_code.decode('utf-8')
        lines = source_code.splitlines()
        total_lines = len(lines)
        # read file pattern and onvert
        #new_source_code = convert_with_pattern_file_excel_by_lines(lines, conversion_rules)
        new_source_code = convert_with_pattern_file_excel_all(source_code, conversion_rules)
        warnings.simplefilter("ignore", category=UserWarning)
        print(f'    ### Total lines code in the file {file}: {total_lines}')
        return new_source_code
    except Exception as e:
        print('Error: ' + str(e))
        logging.info("[process_code_folder] ERROR: " + str(e))
        return jsonify({'status': 'error', 'message': str(e)})


def process_code_file(file, conversion_rules):
    try:
        source_code = file.read()
        if isinstance(source_code, bytes):
            source_code = source_code.decode('utf-8')
        lines = source_code.splitlines()
        total_lines = len(lines)
        # read file pattern and onvert
        # new_source_code = convert_with_pattern_file_excel_by_lines(lines, conversion_rules)
        new_source_code = convert_with_pattern_file_excel_all(source_code, conversion_rules)
        warnings.simplefilter("ignore", category=UserWarning)
        print(f'    ### Total lines code in the file {file}: {total_lines}')
        return new_source_code
    except Exception as e:
        print('Error: ' + str(e))
        logging.info("[process_code_file] ERROR: " + str(e))
        return jsonify({'status': 'error', 'message': str(e)})
